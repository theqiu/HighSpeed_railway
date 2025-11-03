"""Utility for populating the wheel pressing analysis workbook.

This script reads raw inspection records from ``data/2024-2025.csv`` and
injects the matching measurements into ``data/轮对压装统计分析表-2025.10.30.xlsx``.

The Excel template uses a two-row header: the first row contains the
presentation labels while the second row stores the detailed measurement
names that are recorded inside the MES export (CSV).  The script treats the
second header row as the lookup key when searching for a wheel's inspection
records.  Columns whose presentation label already denotes an averaged or
derived metric (``*均``/``*平均``) are skipped to avoid overwriting workbook
formulas.

Usage::

    python fill_pressing_report.py [--csv 自定义CSV路径] [--workbook 模板路径]

When a wheel number already exists inside the workbook the values are merged
into the same row instead of creating duplicates.  Newly discovered wheels are
appended to the end and continue the existing sequence numbering.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
import logging
import re
import argparse
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Tuple, TYPE_CHECKING, Set

import pandas as pd

try:  # pragma: no cover - optional import guard
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore

if TYPE_CHECKING:  # pragma: no cover - typing only
    from openpyxl.cell import Cell
    from openpyxl.worksheet.worksheet import Worksheet
else:  # pragma: no cover
    Cell = Any
    Worksheet = Any


BASE_DIR = Path(__file__).resolve().parent
CSV_PATH = BASE_DIR / "data" / "2024-2025.csv"
WORKBOOK_PATH = BASE_DIR / "data" / "轮对压装统计分析表-2025.10.30.xlsx"

# Measurements often contain decimal places.  Longer numeric strings (e.g.
# 20240117110418) should remain textual, so the matcher only converts values
# that either include a decimal point or are short integer tokens.
NUMERIC_PATTERN = re.compile(r"^-?\d+(?:\.\d+)?$")

# Columns whose labels include these tokens are assumed to rely on workbook
# formulas (averages or summary statistics) and are therefore skipped.
SKIP_TOKENS = ("均", "平均", "过盈量")

ENVIRONMENT_DETAIL_KEYS_RAW: Set[str] = {
    "压装设备编号",
    "压装作业设备",
    "车轴压装同温时长",
    "同温时间",
    "环境温度",
    "温度",
    "环境相对湿度",
    "湿度",
}

ENVIRONMENT_NORMALISED: Set[str]

PEOPLE_CATEGORY_BY_DISPLAY = {
    "轮座测量人": "wheel_seat",
    "轮毂孔测量人": "hub_hole",
    "轮位及内侧距测量人": "position",
    "端跳测量人": "runout",
}

PEOPLE_CATEGORY_BY_DETAIL_RAW = {
    "作业人员-D24A3A3B": "wheel_seat",
    "作业人员-D24A4ED6": "hub_hole",
    "作业人员-D24A7CC2": "position",
    "作业人员-CAF43155": "runout",
}


def normalise_key(value: str) -> str:
    """Return a normalised token for matching template labels to qcitem rows."""

    if value is None:
        return ""
    # Remove whitespace, punctuation commonly used in headers, and harmonise case.
    cleaned = re.sub(r"[\s:：;；\-_/\\]", "", str(value))
    return cleaned.upper()


ENVIRONMENT_NORMALISED = {normalise_key(key) for key in ENVIRONMENT_DETAIL_KEYS_RAW}

PEOPLE_CATEGORY_BY_DETAIL = {
    normalise_key(key): value for key, value in PEOPLE_CATEGORY_BY_DETAIL_RAW.items()
}


@dataclass
class ColumnDescriptor:
    """Metadata for a target Excel column."""

    index: int
    display_name: Optional[str]
    detail_name: Optional[str]

    @property
    def should_skip(self) -> bool:
        display = self.display_name or ""
        detail = self.detail_name or ""
        return any(token in display for token in SKIP_TOKENS) or any(
            token in detail for token in SKIP_TOKENS
        )


WHEEL_DETAIL_KEYS = {normalise_key("轮对号"), normalise_key("轮对编号")}


@dataclass
class Record:
    """A simplified row extracted from the MES CSV."""

    inputvalue: str
    creator: str
    qcitem: str
    opno: str


LOGGER = logging.getLogger(__name__)


def configure_logging(verbose: bool = False) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(levelname)s %(message)s", force=True)
    if verbose:
        LOGGER.debug("Verbose logging enabled")


def categorise_detail(detail: str, display: str = "") -> Optional[str]:
    token = detail or display
    if not token:
        return None
    norm = normalise_key(token)
    if "轮座" in token or "轮座" in display or "LUNZUO" in norm:
        return "wheel_seat"
    if "毂孔" in token or "毂孔" in display or "GUKONG" in norm:
        return "hub_hole"
    if any(keyword in token for keyword in ("轮位", "内侧距")) or any(
        keyword in display for keyword in ("轮位", "内侧距")
    ) or any(key in norm for key in ("LUNWEI", "NEICEJU")):
        return "position"
    if "端跳" in token or "端跳" in display or "DUANTIAO" in norm:
        return "runout"
    return None


def is_meaningful_measurement(detail: str, display: str = "") -> bool:
    norm = normalise_key(detail)
    if norm in WHEEL_DETAIL_KEYS or norm in ENVIRONMENT_NORMALISED:
        return False
    text = detail or display or ""
    keywords = ("轮座", "毂孔", "轮位", "内侧距", "端跳", "压装力", "圆柱度", "轮辋", "轮缘")
    return any(keyword in text for keyword in keywords)


def load_records(csv_path: Path) -> pd.DataFrame:
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV source not found: {csv_path}")

    df = pd.read_csv(csv_path, dtype=str, keep_default_na=False)
    # Normalise whitespace for downstream lookups.
    for column in ("xb005", "opno", "qcitem", "qcorder", "inputvalue", "creator"):
        if column in df.columns:
            df[column] = df[column].astype(str).str.strip()

    if "xb005" not in df.columns:
        raise ValueError("CSV缺少 'xb005' 列，无法按轮对号分组。")

    return df


def iterate_wheels(df: pd.DataFrame) -> Iterator[Tuple[str, pd.DataFrame]]:
    grouped = df.groupby("xb005", sort=True)
    for wheel_id, group in grouped:
        yield wheel_id, group.sort_values("qcdate") if "qcdate" in group.columns else group


def build_lookup(group: pd.DataFrame) -> Tuple[Dict[str, List[Record]], Dict[str, List[Record]]]:
    by_qcitem: Dict[str, List[Record]] = defaultdict(list)
    by_opno: Dict[str, List[Record]] = defaultdict(list)

    for _, row in group.iterrows():
        record = Record(
            inputvalue=str(row.get("inputvalue", "")).strip(),
            creator=str(row.get("creator", "")).strip(),
            qcitem=str(row.get("qcitem", "")).strip(),
            opno=str(row.get("opno", "")).strip(),
        )
        if record.qcitem:
            by_qcitem[record.qcitem].append(record)
            norm_qcitem = normalise_key(record.qcitem)
            if norm_qcitem and norm_qcitem != record.qcitem:
                by_qcitem[norm_qcitem].append(record)
        if record.opno:
            by_opno[record.opno].append(record)
            norm_opno = normalise_key(record.opno)
            if norm_opno and norm_opno != record.opno:
                by_opno[norm_opno].append(record)

    return by_qcitem, by_opno


def normalise_value(raw: str) -> Optional[object]:
    if raw is None:
        return None
    value = raw.strip()
    if not value:
        return None

    if NUMERIC_PATTERN.match(value) and not (value.isdigit() and len(value) >= 8 and "." not in value):
        if "." in value:
            return float(value)
        try:
            return int(value)
        except ValueError:
            # Fall back to string if it doesn't fit into Python's int range.
            return value

    return value


def resolve_field(
    detail_name: str,
    by_qcitem: Dict[str, List[Record]],
    by_opno: Dict[str, List[Record]],
    *,
    wheel_id: Optional[str] = None,
) -> Tuple[Optional[object], Optional[Record]]:
    if not detail_name:
        return None, None

    detail_name = detail_name.strip()
    candidates = by_qcitem.get(detail_name)
    if not candidates:
        normalised = normalise_key(detail_name)
        if normalised and normalised != detail_name:
            candidates = by_qcitem.get(normalised)
        if not candidates:
            candidates = by_opno.get(detail_name)
        if not candidates and normalised and normalised != detail_name:
            candidates = by_opno.get(normalised)
        if not candidates:
            LOGGER.debug(
                "[%s] 未找到匹配的记录: %s",
                wheel_id or "?",
                detail_name,
            )
            return None, None

    record = candidates[0]
    if any(token in detail_name for token in ("作业人员", "测量人")):
        value = record.creator or record.inputvalue
    elif detail_name in {"轮对号", "轮对编号"}:
        # `轮对号` columns simply echo the wheel identifier; this branch allows
        # templates that redundantly store the name in the detail row.
        value = record.qcitem or record.inputvalue
    else:
        value = record.inputvalue or record.creator

    normalised = normalise_value(value)
    LOGGER.debug(
        "[%s] 命中 %s -> %s",
        wheel_id or "?",
        detail_name,
        normalised,
    )
    return normalised, record


def load_workbook_columns(sheet: Worksheet) -> List[ColumnDescriptor]:
    columns: List[ColumnDescriptor] = []
    for idx in range(1, sheet.max_column + 1):
        display = sheet.cell(row=1, column=idx).value
        detail = sheet.cell(row=2, column=idx).value
        if display is None and detail is None:
            continue
        columns.append(ColumnDescriptor(index=idx, display_name=display, detail_name=detail))
    return columns


def scan_existing_rows(
    sheet: Worksheet, columns: Iterable[ColumnDescriptor]
) -> Tuple[Dict[str, int], int, int, int]:
    wheel_columns = [
        column.index
        for column in columns
        if (column.detail_name and normalise_key(column.detail_name) in WHEEL_DETAIL_KEYS)
        or (column.display_name and "轮对" in column.display_name)
    ]
    if not wheel_columns:
        wheel_columns = [column.index for column in columns if column.index == 2]

    existing: Dict[str, int] = {}
    last_seq = 0
    last_data_row = 2

    max_column = sheet.max_column
    for row in range(3, sheet.max_row + 1):
        row_values = [sheet.cell(row=row, column=col).value for col in range(1, max_column + 1)]
        if all(value in (None, "") for value in row_values):
            continue
        last_data_row = max(last_data_row, row)

        seq_value = row_values[0]
        try:
            if isinstance(seq_value, (int, float)):
                last_seq = max(last_seq, int(seq_value))
            elif seq_value not in (None, ""):
                last_seq = max(last_seq, int(str(seq_value)))
        except (TypeError, ValueError):
            pass

        for idx in wheel_columns:
            cell_value = sheet.cell(row=row, column=idx).value
            if cell_value in (None, ""):
                continue
            key = normalise_key(str(cell_value).strip())
            existing.setdefault(key, row)

    next_row = last_data_row + 1
    return existing, next_row, last_seq, last_data_row


def extend_formulas(
    sheet: Worksheet,
    columns: Iterable[ColumnDescriptor],
    *,
    start_row: int,
    end_row: int,
) -> None:
    if end_row < start_row:
        return

    for column in columns:
        if column.index == 1:
            continue

        source_formula = None
        source_row = None
        for row in range(3, end_row + 1):
            cell_value = sheet.cell(row=row, column=column.index).value
            if isinstance(cell_value, str) and cell_value.startswith("="):
                source_formula = cell_value
                source_row = row
                break

        if not source_formula:
            continue

        start_fill = max((source_row or 3) + 1, start_row)
        for row in range(start_fill, end_row + 1):
            target_cell = sheet.cell(row=row, column=column.index)
            if target_cell.value in (None, ""):
                target_cell.value = source_formula


def write_rows(
    sheet: Worksheet,
    columns: List[ColumnDescriptor],
    wheel_data: Iterator[Tuple[str, pd.DataFrame]],
) -> Tuple[int, int]:
    existing_rows, next_row, last_seq, last_data_row = scan_existing_rows(sheet, columns)
    written = 0
    max_row_used = last_data_row

    for wheel_id, group in wheel_data:
        wheel_id = str(wheel_id).strip()
        if not wheel_id:
            continue

        wheel_key = normalise_key(wheel_id)
        by_qcitem, by_opno = build_lookup(group)
        LOGGER.info("处理轮对 %s，记录数 %s", wheel_id, len(group))

        assignments: Dict[int, object] = {}
        creators_by_category: Dict[str, str] = {}
        people_columns: List[Tuple[ColumnDescriptor, str]] = []
        has_meaningful = False

        for column in columns:
            if column.index == 1:
                continue

            if column.should_skip:
                continue

            detail = column.detail_name or ""
            display = column.display_name or ""

            people_category: Optional[str] = None
            if display in PEOPLE_CATEGORY_BY_DISPLAY:
                people_category = PEOPLE_CATEGORY_BY_DISPLAY[display]
            else:
                norm_detail = normalise_key(detail)
                people_category = PEOPLE_CATEGORY_BY_DETAIL.get(norm_detail)

            if people_category:
                people_columns.append((column, people_category))
                continue

            lookup_key = detail or display
            value: Optional[object] = None
            record: Optional[Record] = None

            if normalise_key(lookup_key) in WHEEL_DETAIL_KEYS or "轮对号" in display:
                value = wheel_id
            else:
                value, record = resolve_field(lookup_key, by_qcitem, by_opno, wheel_id=wheel_id)

            if value is None:
                LOGGER.debug(
                    "[%s] 列 %s/%s 未找到匹配值",
                    wheel_id,
                    column.display_name,
                    detail,
                )
                continue

            assignments[column.index] = value

            if is_meaningful_measurement(detail, display):
                has_meaningful = True

            category = categorise_detail(detail, display)
            if record and record.creator and category:
                creators_by_category.setdefault(category, record.creator)

        for column, category in people_columns:
            creator_value = creators_by_category.get(category)
            if creator_value:
                assignments[column.index] = creator_value

        if not has_meaningful:
            LOGGER.info("轮对 %s 缺少有效测量数据，跳过写入。", wheel_id)
            continue

        target_row = existing_rows.get(wheel_key)
        new_row = False
        if target_row is None:
            target_row = next_row
            next_row += 1
            last_seq += 1
            new_row = True
            assignments[1] = last_seq
        else:
            if sheet.cell(row=target_row, column=1).value in (None, ""):
                last_seq += 1
                assignments[1] = last_seq

        for column in columns:
            if column.index == 1 and column.index not in assignments:
                continue
            if column.index not in assignments:
                continue

            value = assignments[column.index]
            target_cell: Cell = sheet.cell(row=target_row, column=column.index)

            if isinstance(target_cell.value, str) and target_cell.value.startswith("="):
                continue

            if (target_cell.value not in (None, "")) and not new_row:
                continue

            target_cell.value = value

        max_row_used = max(max_row_used, target_row)
        existing_rows[wheel_key] = target_row
        written += 1

    return written, max_row_used


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fill the wheel pressing workbook")
    parser.add_argument(
        "--csv",
        type=Path,
        default=CSV_PATH,
        help="Path to the MES CSV export (default: data/2024-2025.csv)",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=WORKBOOK_PATH,
        help="Path to the Excel workbook to populate",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Optional worksheet name; defaults to the active sheet",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable verbose logging for debugging column matches",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    configure_logging(args.verbose)

    if load_workbook is None:  # pragma: no cover - handled during runtime
        raise RuntimeError("openpyxl 未安装，无法读写工作簿。")

    df = load_records(args.csv)
    wb = load_workbook(args.workbook)
    sheet = wb[args.sheet] if args.sheet else wb.active
    columns = load_workbook_columns(sheet)

    if not columns:
        raise ValueError("目标工作簿缺少表头，无法写入数据。")

    total_written, max_row_used = write_rows(sheet, columns, iterate_wheels(df))
    extend_formulas(sheet, columns, start_row=3, end_row=max_row_used)

    if total_written == 0:
        logging.warning("未写入任何轮对数据，请检查模板列名与CSV是否匹配。")
    else:
        logging.info("成功写入 %s 条轮对记录。", total_written)

    wb.save(args.workbook)
    logging.info("结果已保存至 %s", args.workbook)


if __name__ == "__main__":
    main()

