from __future__ import annotations

from pathlib import Path
import tempfile
import unittest

import pandas as pd

try:  # pragma: no cover - optional dependency guard for test environment
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None

from fill_pressing_report import (
    iterate_wheels,
    load_records,
    load_workbook_columns,
    normalise_key,
    write_rows,
)


@unittest.skipIf(Workbook is None or load_workbook is None, "openpyxl not installed")
class FillPressingReportTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmpdir.cleanup)
        self.base = Path(self.tmpdir.name)

    def create_csv(self) -> Path:
        rows = [
            {
                "xb005": "109M3C09877",
                "qcitem": "左轮座直径A1",
                "inputvalue": "500.11",
                "creator": "MEASURE-A",
                "qcdate": "20240101120000",
                "opno": "轴盘压装",
            },
            {
                "xb005": "109M3C09877",
                "qcitem": "左轮座直径A2",
                "inputvalue": "500.22",
                "creator": "MEASURE-A",
                "qcdate": "20240101120500",
                "opno": "轴盘压装",
            },
            {
                "xb005": "109M3C09877",
                "qcitem": "右轮座直径B3",
                "inputvalue": "601.33",
                "creator": "MEASURE-B",
                "qcdate": "20240101121000",
                "opno": "轴盘压装",
            },
            {
                "xb005": "109M3C09877",
                "qcitem": "作业人员-D24A3A3B",
                "inputvalue": "",
                "creator": "WORKER-1",
                "qcdate": "20240101122000",
                "opno": "轮对压装",
            },
            {
                "xb005": "109M3C09877",
                "qcitem": "轮对号",
                "inputvalue": "109M3C09877",
                "creator": "",
                "qcdate": "20240101122500",
                "opno": "轮对压装",
            },
        ]
        df = pd.DataFrame(rows)
        csv_path = self.base / "source.csv"
        df.to_csv(csv_path, index=False)
        return csv_path

    def create_workbook(self) -> Path:
        wb = Workbook()
        ws = wb.active
        headers = [
            ("序号", "序号"),
            ("轮对号", "轮对号"),
            ("作业人员", "作业人员-D24A3A3B"),
            ("左轮座A1", "左轮座直径A1"),
            ("左轮座A均", "左轮座直径A"),
            ("右轮座B3", "右轮座直径B3"),
            ("左轮过盈量", "左轮配合过盈量"),
        ]
        for col, (display, detail) in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=display)
            ws.cell(row=2, column=col, value=detail)
        wb_path = self.base / "template.xlsx"
        wb.save(wb_path)
        return wb_path

    def test_normalise_key_removes_spacing(self) -> None:
        self.assertEqual(normalise_key(" 左 轮 座 直 径A1"), "左轮座直径A1".upper())

    def test_write_rows_populates_expected_cells(self) -> None:
        csv_path = self.create_csv()
        wb_path = self.create_workbook()

        df = load_records(csv_path)
        wb = load_workbook(wb_path)
        sheet = wb.active
        columns = load_workbook_columns(sheet)

        written = write_rows(sheet, columns, iterate_wheels(df))

        self.assertEqual(written, 1)
        # First data row should start at row 3 with sequence number 1.
        self.assertEqual(sheet.cell(row=3, column=1).value, 1)
        self.assertEqual(sheet.cell(row=3, column=2).value, "109M3C09877")
        self.assertEqual(sheet.cell(row=3, column=3).value, "WORKER-1")
        self.assertAlmostEqual(sheet.cell(row=3, column=4).value, 500.11)
        # Average column should retain formula/None (skipped because of 均 token).
        self.assertIsNone(sheet.cell(row=3, column=5).value)
        self.assertAlmostEqual(sheet.cell(row=3, column=6).value, 601.33)
        # Column containing 过盈量 should be skipped entirely.
        self.assertIsNone(sheet.cell(row=3, column=7).value)


if __name__ == "__main__":
    unittest.main()
